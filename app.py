"""
Menu Digitalizer - foodora

Selgeren laster opp en meny (PDF, Word, Excel eller bilde).
Verktoyet ekstraherer rettene, normaliserer tekst, utleder allergener,
og viser alt i et redigerbart grid. Selgeren retter ved behov og laster
ned en Excel-fil i MDS-formatet, navngitt <Vendor>_<GRID>.xlsx.
"""

import io
import re

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from extraction import extract_menu
from rules import detect_allergens, to_title_case, to_sentence_case

# --- MDS-kolonner. Speiler malen fra MDS sitt CSV-output. -------------------
# Allergens er den nye kolonnen som ikke finnes i standard MDS.
MDS_COLUMNS = [
    "Title_en_NO", "Title_en_GB", "Title_zh_HK", "Title_en_US",
    "Description_en_NO", "Description_en_GB", "Description_zh_HK",
    "Description_en_US", "Description_type", "Category_ID",
    "Pre - packed", "Active", "Image_URL", "VAT_ID",
    "Variation_title_en_NO", "Variation_title_en_GB",
    "Variation_title_zh_HK", "Variation_title_en_US",
    "Price", "Remote_Code", "Container_Charge", "Choice_Groups_IDs",
    "Allergens",  # <-- ny kolonne
]

# --- foodora / Delivery Hero fargepalett ------------------------------------
FOODORA_PINK = "#FF1F62"
FOODORA_PINK_DARK = "#D81B54"
INK = "#1A1A2E"
SOFT_BG = "#FFF5F8"
BORDER = "#FFD6E2"

# Marked -> sprakkode. Kun NO aktivt naa; lett aa utvide senere.
MARKETS = {
    "Norge": "NO",
    # "Storbritannia": "GB",
    # "USA": "US",
}

st.set_page_config(page_title="Menu Digitalizer - foodora",
                   page_icon="\U0001F37D", layout="wide")

# --- Stil: gjor appen foodora-rosa i stedet for grastandard -----------------
st.markdown(f"""
<style>
  .stApp {{ background-color: #FFFFFF; }}
  h1, h2, h3 {{ color: {INK}; }}
  /* primaerknapper */
  .stButton > button[kind="primary"],
  .stDownloadButton > button {{
      background-color: {FOODORA_PINK};
      color: #FFFFFF;
      border: none;
      border-radius: 8px;
      font-weight: 600;
  }}
  .stButton > button[kind="primary"]:hover,
  .stDownloadButton > button:hover {{
      background-color: {FOODORA_PINK_DARK};
      color: #FFFFFF;
  }}
  /* sidebar */
  section[data-testid="stSidebar"] {{
      background-color: {SOFT_BG};
      border-right: 1px solid {BORDER};
  }}
  /* metrikk-kort */
  div[data-testid="stMetric"] {{
      background-color: {SOFT_BG};
      border: 1px solid {BORDER};
      border-radius: 10px;
      padding: 12px 16px;
  }}
  /* topp-banner */
  .md-banner {{
      background: linear-gradient(135deg, {FOODORA_PINK} 0%, {FOODORA_PINK_DARK} 100%);
      color: #FFFFFF;
      padding: 22px 28px;
      border-radius: 14px;
      margin-bottom: 18px;
      display: flex;
      align-items: center;
      gap: 22px;
  }}
  .md-banner h1 {{ color: #FFFFFF; margin: 0; font-size: 30px; }}
  .md-banner p  {{ color: #FFE3EC; margin: 6px 0 0 0; font-size: 15px; }}
  .md-logo {{
      background: #FFFFFF;
      border-radius: 12px;
      padding: 14px 18px;
      flex-shrink: 0;
      display: flex;
      align-items: center;
  }}
  .md-logo img {{ height: 56px; width: auto; display: block; }}
</style>
""", unsafe_allow_html=True)


# ---------------------------------------------------------------------------
# Hjelpefunksjoner
# ---------------------------------------------------------------------------

def items_to_dataframe(items):
    """Gjor raa ekstraksjon om til redigerbart grid med normalisering."""
    rows = []
    for it in items:
        title = to_title_case(str(it.get("title", "")).strip())
        desc = to_sentence_case(str(it.get("description", "")).strip())
        price = it.get("price")
        rows.append({
            "Tittel": title,
            "Beskrivelse": desc,
            "Variant": str(it.get("variation", "")).strip(),
            "Pris (NOK)": price if price is not None else 0.0,
            "Kategori": str(it.get("category", "")).strip(),
            "Allergener": detect_allergens(desc),
        })
    return pd.DataFrame(rows, columns=[
        "Tittel", "Beskrivelse", "Variant", "Pris (NOK)",
        "Kategori", "Allergener",
    ])


def safe_filename_part(text):
    """Gjor en tekst trygg som del av et filnavn."""
    cleaned = re.sub(r"[^\w\-]+", "_", text.strip())
    return cleaned.strip("_")


def build_export_filename(vendor, grid):
    """Bygg filnavn etter MDS-konvensjonen <Vendor>_<GRID>.xlsx."""
    v = safe_filename_part(vendor) or "Vendor"
    g = safe_filename_part(grid) or "GRID"
    return f"{v}_{g}.xlsx"


def build_mds_excel(df, market_lang="NO"):
    """Bygg en Excel-fil i MDS-formatet fra det redigerte gridet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Draft Menu - MDS"

    ws.append(MDS_COLUMNS)
    header_fill = PatternFill("solid", start_color=FOODORA_PINK.lstrip("#"))
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    title_key = f"Title_en_{market_lang}"
    desc_key = f"Description_en_{market_lang}"
    var_key = f"Variation_title_en_{market_lang}"

    for _, r in df.iterrows():
        record = {c: "" for c in MDS_COLUMNS}
        record[title_key] = r["Tittel"]
        record[desc_key] = r["Beskrivelse"]
        record[var_key] = r["Variant"]
        record["Description_type"] = "VENDOR"
        record["Pre - packed"] = "FALSE"
        record["Active"] = "TRUE"
        record["Price"] = r["Pris (NOK)"]
        record["Allergens"] = r["Allergener"]
        ws.append([record[c] for c in MDS_COLUMNS])

    for col_idx, name in enumerate(MDS_COLUMNS, start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        if name.startswith("Description"):
            ws.column_dimensions[letter].width = 45
        elif name.startswith("Title") or name == "Allergens":
            ws.column_dimensions[letter].width = 28
        else:
            ws.column_dimensions[letter].width = 16

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------

def _logo_data_uri():
    """Les Delivery Hero-logoen som base64 slik at den kan ligge i banneret."""
    import base64
    import os
    path = os.path.join(os.path.dirname(__file__), "dh_logo.png")
    try:
        with open(path, "rb") as f:
            b64 = base64.standard_b64encode(f.read()).decode("utf-8")
        return f"data:image/png;base64,{b64}"
    except FileNotFoundError:
        return None


_logo = _logo_data_uri()
_logo_html = (
    f'<div class="md-logo"><img src="{_logo}" alt="Delivery Hero"/></div>'
    if _logo else ""
)

st.markdown(f"""
<div class="md-banner">
  {_logo_html}
  <div>
    <h1>\U0001F37D Menu Digitalizer</h1>
    <p>Last opp en meny &rarr; verkt&oslash;yet strukturerer den &rarr;
       rediger og last ned i MDS-format.</p>
  </div>
</div>
""", unsafe_allow_html=True)

with st.sidebar:
    st.header("Innstillinger")

    market_name = st.selectbox(
        "Marked",
        options=list(MARKETS.keys()),
        help="Bestemmer hvilken Title-/Description-kolonne i MDS-malen "
             "som fylles ut. Flere markeder kan legges til senere.",
    )
    market = MARKETS[market_name]

    st.divider()
    st.subheader("Vendor")
    vendor_name = st.text_input(
        "Vendornavn",
        placeholder="f.eks. Randis Gatekjokken",
        help="Brukes i filnavnet paa den nedlastede Excel-fila.",
    )
    grid_id = st.text_input(
        "GRID-id",
        placeholder="f.eks. a4b2",
        help="Vendor-id paa DH. Brukes i filnavnet etter "
             "MDS-konvensjonen <Vendor>_<GRID>.xlsx.",
    )

    st.divider()
    st.caption(
        "Allergener utledes fra ingrediensene i beskrivelsen og er "
        "alltid merket *antatt \u2013 bekreft*. Selgeren m\u00e5 verifisere "
        "mot vendoren."
    )

# API-nokkelen leses fra Streamlit Secrets - aldri vist i UI.
try:
    api_key = st.secrets.get("ANTHROPIC_API_KEY", "")
except Exception:
    api_key = ""

if "menu_df" not in st.session_state:
    st.session_state.menu_df = None

uploaded = st.file_uploader(
    "Last opp meny",
    type=["pdf", "docx", "xlsx", "xls", "jpg", "jpeg", "png"],
    help="PDF, Word, Excel eller bilde av menyen.",
)
analyze_file = st.button("Analyser meny", type="primary",
                         disabled=uploaded is None)

st.caption(
    "\U0001F4A1 Meny p\u00e5 nett? \u00c5pne siden i nettleseren, scroll "
    "helt til bunnen sl\u00e5 hele menyen er lastet, og lagre siden som "
    "PDF (Cmd+P \u2192 Lagre som PDF). Last s\u00e5 opp PDF-en her."
)

# --- Filopplasting -----------------------------------------------------------
if analyze_file and uploaded is not None:
    with st.spinner("Analyserer menyen \u2026"):
        try:
            items = extract_menu(
                uploaded.getvalue(), uploaded.name, api_key or None)
            if not items:
                st.warning("Fant ingen menyelementer i filen.")
            else:
                st.session_state.menu_df = items_to_dataframe(items)
                st.success(f"Hentet ut {len(items)} element(er). "
                           "Rediger ved behov under.")
        except ValueError as e:
            st.error(str(e))
        except Exception as e:
            st.error(f"Noe gikk galt under analysen: {e}")

# ---------------------------------------------------------------------------
# Redigerbart grid
# ---------------------------------------------------------------------------

if st.session_state.menu_df is not None:
    st.subheader("Rediger menyen")
    st.caption("Klikk i cellene for \u00e5 rette. Tomme priser vises som 0 \u2013 "
               "fyll inn riktig verdi.")

    edited = st.data_editor(
        st.session_state.menu_df,
        use_container_width=True,
        num_rows="dynamic",
        column_config={
            "Tittel": st.column_config.TextColumn(width="medium"),
            "Beskrivelse": st.column_config.TextColumn(width="large"),
            "Variant": st.column_config.TextColumn(width="small"),
            "Pris (NOK)": st.column_config.NumberColumn(
                format="%.0f", min_value=0),
            "Kategori": st.column_config.TextColumn(width="small"),
            "Allergener": st.column_config.TextColumn(width="medium"),
        },
        key="editor",
    )
    st.session_state.menu_df = edited

    missing_price = (edited["Pris (NOK)"] == 0).sum()
    needs_check = edited["Allergener"].str.contains(
        "Sjekk med vendor", na=False).sum()
    c1, c2, c3 = st.columns(3)
    c1.metric("Retter", len(edited))
    c2.metric("Mangler pris", int(missing_price))
    c3.metric("Allergener \u00e5 sjekke", int(needs_check))

    if missing_price:
        st.warning(f"{missing_price} rett(er) har ingen pris \u2013 "
                   "fyll inn f\u00f8r eksport.")

    st.divider()

    export_name = build_export_filename(vendor_name, grid_id)
    if not vendor_name or not grid_id:
        st.info(f"Fyll inn vendornavn og GRID-id i sidepanelet for "
                f"riktig filnavn. Nåværende: **{export_name}**")
    else:
        st.caption(f"Fila lastes ned som: **{export_name}**")

    excel_buf = build_mds_excel(edited, market_lang=market)
    st.download_button(
        "\u2b07\ufe0f Last ned MDS-Excel",
        data=excel_buf,
        file_name=export_name,
        mime="application/vnd.openxmlformats-officedocument."
             "spreadsheetml.sheet",
        type="primary",
    )
